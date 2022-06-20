# -*- coding: utf-8 -*-

import json
import os
import time

import openpyxl as xl
import psutil
import requests


def timeit():
    return time.time()


def memoit():
    process = psutil.Process(os.getpid())
    in_mbs = process.memory_info().rss / 1024 ** 2
    return in_mbs


def run():
    url = f'{base_url}/user/mass/upload'
    wb = xl.load_workbook('users.xlsx')
    ws = wb.active
    print('Loaded file users.xlsx ...')

    data = [
        {
            'name': f'{ws.cell(row, 1).value.strip()} {ws.cell(row, 2).value.strip()}',
            'login': ws.cell(row, 3).value.strip(),
            'role': ws.cell(row, 4).value.strip(),
            'department': ws.cell(row, 6).value.strip(),
        } for row in range(2, ws.max_row + 1)
    ]

    print('Done loading data for file users.xlsx ...')
    headers = {'Content-Type': 'application/json'}
    raw['params'] = {'payload': data}
    print(json.dumps(data, indent=2))
    response = requests.post(url=url, data=json.dumps(raw), headers=headers)
    print(response.text)


if __name__ == '__main__':
    base_url = os.environ.get('BASE_URL', 'http://localhost:8015')
    headers = {'Content-Type': 'application/json'}
    raw = {"jsonrpc": "2.0", "params": False, 'id': None}

    start_time = timeit()
    mem_before = memoit()

    result = run()

    mem_after = memoit()
    end_time = timeit()
    print(f'Process finished in {end_time-start_time}s consuming {mem_after-mem_before}mbs of memory')
