from string import ascii_uppercase as ascp
from datetime import date
import tempfile
import base64
import logging
from odoo import models, _, fields, api
import os
import openpyxl as xl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

_logger = logging.getLogger(__name__)


class OpResultXlsx(models.TransientModel):
    _name = 'op.result.xlsx'

    exam_session = fields.Many2one('op.exam.session', string='Exam Session')

    school_year = fields.Char(string="School Year",
                              required=True, default=str(date.today().year))
    school_term = fields.Many2one('op.course', string='School Term')
    school_class = fields.Selection(
        selection=[('0', 'Kindergatten'),
                   ('1', 'Grade One'), ('2',
                                        'Grade Two'),
                   ('3', 'Grade Three'), ('4',
                                          'Grade Four'),
                   ('5', 'Class Five'), ('6',
                                         'Class Six'),
                   ('7', 'Class Seven'), ('8', 'Class Eight')], string='School Class', compute='_compute_school_class')

    @api.depends('school_term')
    def _compute_school_class(self):
        for rec in self:
            rec.school_class = rec.school_term.class_number

    def get_results_xlsx_report(self):
        subjects = self.env['op.subject'].search([]).mapped('name')
        data = {}
        exam_name = ''
        codes = ['students']
        subject_totals = {}

        exams = self.env['op.exam.session'].search(
            [('id', '=', self.exam_session.id)]).filtered(lambda x: x.start_date.year == int(self.school_year))

        for exam in exams:
            exam_name = exam.name
            for line in exam.exam_ids:
                codes.append(line.subject_id.code)

                for student in line.attendees_line:
                    if student.student_id.name in data:
                        data[student.student_id.name][line.subject_id.code] = student.marks
                    else:
                        data[student.student_id.name] = {}

                        data[student.student_id.name].setdefault(
                            line.subject_id.code, student.marks)

        for student in data:
            total = average = 0
            for exam in data[student]:
                total += data[student][exam]

            average = round(total / len(data[student]), 2)
            data[student]['average'] = average
            data[student]['total'] = total

        for student in data:
            for exam in data[student]:
                if exam in subject_totals:
                    subject_totals[exam] += data[student][exam]
                else:
                    subject_totals[exam] = data[student][exam]

        codes.append('average')
        codes.append('total')

        _logger.error(f'-----------------------{data}------{codes}')

        return data, codes, exam_name, subject_totals

    def create_results_xlsx_file(self):
        data, codes, exam_name, subject_totals = self.get_results_xlsx_report()

        wb = xl.Workbook()
        ws = wb.active
        big_font = Font(name='Arial', size=13, bold=True)
        mid_font = Font(name='Arial', size=11, bold=True)
        normal_font = Font(name='Arial', size=10)
        center_align = Alignment(horizontal="center", vertical="center")

        # img = xl.drawing.image.Image(
        #     f'{os.getcwd()}custom-addons/oo_openeducat_parent_web/static/img/logo.png')

        # ws.add_image(img, 'A1')
        ws.merge_cells('C4:I4')
        ws['C4'] = exam_name
        ws['C4'].font = big_font
        ws['C4'].alignment = center_align

        for cell, title in zip(list(ascp[:len(codes)]), codes):
            ws[f'{cell}6'] = title.capitalize()
            ws[f'{cell}6'].font = big_font

        fr = 7
        for student in data:
            ws[f'A{fr}'] = student

            for i, code in enumerate(codes[1:], start=2):
                letter = get_column_letter(i)
                ws[f'{letter}{fr}'] = data[student][code]
                ws[f'{letter}{fr}'].font = normal_font

                if i in (codes[-1], codes[-2]):
                    ws[f'{letter}{fr}'].font = mid_font

            fr += 1

        for subject in subject_totals:
            ws[f'A{fr}'] = 'Totals'
            ws[f'A{fr}'].font = mid_font
            ws[f'A{fr}'].alignment = center_align

            for i, code in enumerate(codes[1:], start=2):
                letter = get_column_letter(i)
                ws[f'{letter}{fr}'] = subject_totals[code]
                ws[f'{letter}{fr}'].font = mid_font

        file_path = self.create_xls_path()
        wb.save(file_path)
        attachment_id = self.save_xlsx(file_path, exam_name)
        return {
            'name': _('Download Results'),
            'context': self.env.context,
            'view_type': 'form',
            'view_mode': 'form',
            'res_model': 'wizard.excel.result',
            'res_id': attachment_id.id,
            'data': None,
            'type': 'ir.actions.act_window',
            'target': 'new'
        }

    def create_xls_path(self):
        _, file_path = tempfile.mkstemp(
            suffix='.xlsx', prefix='school_results')

        return file_path

    def save_xlsx(self, file_path, exam_name):
        try:
            with open(file_path, 'rb') as f:
                attachment_id = self.env['wizard.excel.result'].create({
                    'name': f'{exam_name}.xlsx',
                    'report': base64.encodestring(f.read())
                })
            os.unlink(file_path)
        except (OSError, IOError):
            _logger.error(
                'Error when trying to remove file %s' % file_path)

        return attachment_id


class WizardExcelReport(models.TransientModel):
    _name = 'wizard.excel.result'
    _description = "Results Excel Report"

    name = fields.Char('File Name', size=64)
    report = fields.Binary('Excel Report', readonly=True)
