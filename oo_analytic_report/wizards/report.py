# -*- coding: utf-8 -*-

from odoo import models, fields, api, _
import openpyxl as XL
import base64
from openpyxl.styles import Alignment, Font, numbers
from openpyxl.utils import get_column_letter
from datetime import date
from collections import OrderedDict as OD
import logging
from odoo.exceptions import ValidationError
import io
import calendar
_logger = logging.getLogger(__name__)


class ReportDownload(models.TransientModel):
    _name = 'excel.wizard'
    _description = 'Download Excel Forms'

    name = fields.Char('File Name', size=64)
    report = fields.Binary('Download File (.xlsx)', readonly=True)


class AnalyticReport(models.TransientModel):
    _name = 'analytic.report.wizard'
    _description = 'Generate analytic account wizards'

    period = fields.Selection(string='Timeframe', selection=[(
        'today', 'Today'), ('month', 'This Month'), ('between', 'Custom Dates')], default='today', required=True)
    date_start = fields.Date(string='Start Date')
    date_end = fields.Date(string='End Date', default=fields.Date.today())

    def _prepare_data(self):
        fetch = []
        today = fields.Date.today()
        AccAnalyticLine = self.env['account.analytic.line']
        if self.period == 'today':
            fetch = AccAnalyticLine.search([('date', '=', today)])
        if self.period == 'month':
            month_start = today.strftime('%Y-%m') + '-01'
            fetch = AccAnalyticLine.search(
                [('date', '>=', month_start), ('date', '<=', today)])

        if self.period == 'between':
            fetch = AccAnalyticLine.search(
                [('date', '>=', self.date_start), ('date', '<=', self.date_end)])

        data = {}
        if fetch:
            for line in fetch:
                data[line.id] = {'sale': [], 'purchase': []}
                vals = {
                    'name': line.name,
                    'quantity': line.quantity,
                    'date': line.date,
                    'partner': line.account_id.partner_id.name,
                    'sale_id': line.account_id.name,
                    'product_id': line.product_id.name,
                    'categ_id': line.product_id.categ_id.name,
                    'tax': line.amount,
                    'amount_untaxed': line.amount,
                    'amount_total': line.amount,
                    'payment_mode': line.move_id.move_id.journal_id.name,
                }
                if line.amount >= 0:
                    data[line.id]['sale'].append(vals)
                else:
                    vals.update({
                        'purchase_rep': '',
                        'supplier': '',
                        'purchase_id': '',
                        'description': '',
                        'quantity': '',
                        'price_unit': 100,
                    })
                    data[line.id]['purchase'].append(vals)

        return data

    def generate_report(self):
        filename = 'Analytic Report'
        wb = XL.Workbook()
        ws = wb.active
        big_font = Font(name='Arial', size=13, bold=True)
        mid_font = Font(name='Arial', size=11, bold=True)
        normal_font = Font(name='Arial', size=10)
        center_align = Alignment(horizontal="center", vertical="center")

        sales = ['Customer', 'Analytic Account', 'Product Category', 'Product', 'Quantity',
                 'Unit Price', 'Exclusive Amount', 'VAT', 'Total Amount', 'Payment Mode']
        purchase = ['Purchase Representative', 'Supplier',
                    'Purchase Order', 'Description', 'Quantity', 'Unit Price', 'Exclusive Amount', 'VAT', 'Total Amount', 'Payment Mode']

        cols = OD()
        cols['sale'] = sales
        cols['purchase'] = purchase

        period_sel = dict(self._fields['period'].selection).get(
            self.period)
        period = period_sel if self.period != 'between' else "{start} To {stop}".format(
            start=self.date_start.strftime('%d %B,%Y'), stop=self.date_end.strftime('%d %B,%Y'))

        ws['A1'] = filename
        ws['A1'].font = big_font
        ws['A2'] = 'Report Period'
        ws['A2'].font = big_font
        ws['B2'] = period
        ws['B2'].font = big_font
        ws.merge_cells('B2:F2')

        fr = 4
        col_index = 1
        for key, value in cols.items():
            for col in value:
                letter = get_column_letter(col_index)
                ws[letter + str(fr)] = col
                ws[letter + str(fr)].font = mid_font
                ws[letter + str(fr)].alignment = center_align
                col_index += 1

            col_index += 1
        fr += 1
        data = self._prepare_data()

        for key in data:
            sale = data['sale']
            purchase = data['purchase']

        buf = io.BytesIO()
        wb.save(buf)
        out = base64.encodestring(buf.getvalue())

        attachment_id = self.env['excel.wizard'].create({
            'name': 'Analytic Report.xlsx',
            'report': out
        })
        return {
            'name': _('Download Analytic Report'),
            'context': self.env.context,
            'view_mode': 'form',
            'res_model': 'excel.wizard',
            'res_id': attachment_id.id,
            'data': None,
            'type': 'ir.actions.act_window',
            'target': 'new'
        }
