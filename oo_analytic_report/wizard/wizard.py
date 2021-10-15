# -*- coding: utf-8 -*-

import base64
import logging
import os
import tempfile
from collections import OrderedDict

import openpyxl as xl
from odoo import api, fields, models, _
from odoo.exceptions import ValidationError
from openpyxl.styles import Alignment, Font, numbers
import io
from openpyxl.utils import get_column_letter

_logger = logging.getLogger(__name__)


class ReportDownload(models.TransientModel):
    _name = 'excel.wizard'
    _description = 'Download Excel Forms'

    name = fields.Char('File Name', size=64)
    report = fields.Binary('Excel Report', readonly=True)


class AnalyticExcelReport(models.TransientModel):
    _name = 'analytic.excel.report'

    report_type = fields.Selection(string='Report Type', selection=[(
        'daily', 'Daily'), ('monthly', 'Monthly')], default='daily', required=True)
    day_of = fields.Date(string='On')
    month_of = fields.Char(string='For Month')

    def _analytic_excel_columns(self):
        od = OrderedDict()
        od['sale'] = ['Analytic Account', 'Product Category', 'Product',
                      'Quantity', 'Unit Price', 'Amount Less Vat',
                      'VAT', 'Total Amount', 'Payment Mode']
        od['purchase'] = ['Purchase Rep', 'Vendor', 'Purchase Order',
                          'Description', 'Quantity', 'Unit Price', 'Amount Less Vat',
                          'VAT', 'Total Amount', 'Payment Mode', 'Balance']

        return od

    @api.multi
    def action_generate_report(self):
        wb = xl.Workbook()
        ws = wb.active
        bfont = Font(name='Arial', bold=True, size=14)
        mfont = Font(name='Arial', bold=True, size=14)
        nfont = Font(name='Arial', size=13)
        centerAlign = Alignment(horizontal="center", vertical="center")
        filename = 'Analytic Report'
        # ws.merge_cells('B2:D2')

        cols = self._analytic_excel_columns()
        fr = 3
        count = 1
        for col in sorted(cols):
            for title in cols[col]:
                letter = get_column_letter(count)
                cell = '{letter}{fr}'.format(letter=letter, fr=fr)
                ws[cell] = title
                ws[cell].font = bfont
                ws[cell].alignment = centerAlign
                count += 1

            count += 1

        fr += 1

        analytics = self.env['account.analytic.line'].search(
            [('date', '=', self.day_of)])

        data = {}
        for line in analytics:
            vals = {
                'account_id': line.account_id.name or '',
                'category_id': line.product_id.categ_id.name or '',
                'product_id': line.product_id.name or '',
                'quantity': line.unit_amount or '',
                'price': abs(line.amount / line.unit_amount) or '',
                'untaxed_amount': line.amount or '',
                'tax': 0 or '',
                'amount': line.amount or 0,
                'payment_mode': line.move_id.journal_id.name or '',

            }
            if line.account_id.name in data:
                if line.amount > 0:
                    data[line.account_id.name]['sale'].append(vals)
                else:
                    data[line.account_id.name]['purchase'].append(vals)
            else:
                if line.amount > 0:
                    data[line.account_id.name]['sale'] = [vals]
                else:
                    data[line.account_id.name]['purchase'] = [vals]

        buf = io.BytesIO()
        wb.save(buf)
        out = base64.encodestring(buf.getvalue())

        attachment_id = self.env['excel.wizard'].create({
            'name': 'Analytic Report.xlsx',
            'report': out
        })
        return {
            'name': _('Download Analytic Report'),
            'context': self.env.context,
            'view_mode': 'form',
            'res_model': 'excel.wizard',
            'res_id': attachment_id.id,
            'data': None,
            'type': 'ir.actions.act_window',
            'target': 'new'
        }
