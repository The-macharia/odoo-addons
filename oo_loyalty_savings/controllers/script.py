# -*- coding: utf-8 -*-
import requests
import openpyxl as xl
import json
import os


def send_loyalty(data):
    url = f'{base_url}/loyalty/upload'
    raw['params'] = data
    payload = json.dumps(raw)

    res = requests.post(url=url, data=payload, headers=headers)
    print(res)


def send_savings(data):
    raw['params'] = data
    payload = json.dumps(raw)

    url = f'{base_url}/savings/upload'
    res = requests.post(url=url, data=payload, headers=headers)
    print(res)


def run(file):
    wb = xl.load_workbook(file)
    loyalty = {}
    savings = {}
    for sheet in wb:
        if sheet.title == 'Savings':
            for row in range(2, sheet.max_row + 1):
                group = sheet.cell(row, 1).value.strip()
                line = {
                    'partner': sheet.cell(row, 2).value,
                    'points': sheet.cell(row, 3).value,
                }
                if savings.get(group):
                    savings[group].append(line)
                else:
                    savings[group] = [line]

        if sheet.title == 'Loyalty':
            for row in range(2, sheet.max_row + 1):
                group = sheet.cell(row, 1).value.strip()
                line = {
                    'partner': sheet.cell(row, 2).value,
                    'points': sheet.cell(row, 3).value,
                }
                if loyalty.get(group):
                    loyalty[group].append(line)
                else:
                    loyalty[group] = [line]
    # print(json.dumps(savings, indent=2))
    # print(json.dumps(loyalty, indent=2))
    send_loyalty(loyalty)
    send_savings(savings)


if __name__ == '__main__':
    headers = {'Content-Type': 'application/json'}
    raw = {"jsonrpc": "2.0", "params": False, 'id': None}

    base_url = os.environ.get('BASE_URL')
    run('file.xlsx')
