# -*- coding: utf-8 -*-
import base64
import logging
import os
import tempfile
from datetime import date, datetime

import openpyxl as xl
from odoo import _, api, fields, models
from odoo.tools.float_utils import float_round
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


BIG_FONT = Font(name='Arial', bold=True, size=15)
MID_FONT = Font(name='Arial', bold=True, size=12)
NORMAL_FONT = Font(name='Arial', size=10)
CENTER_ALIGN = Alignment(horizontal='left', vertical='center')

DATE_FORMAT = '%d/%m/%Y'
DATETIME_FORMAT = '%d/%m/%Y %H:%M:%S'

_logger = logging.getLogger(__name__)

BY_DATE_COLS = ['Customer', 'Route', 'Sales', 'Paid', 'Balance', 'Cumulative Bal']
ANALYSIS_COLS = ['Date', 'Dimension', 'Customer', 'Product',
                 'QTY (pcs)', 'QTY (crates)', 'Selling Price', 'Unit Cost', 'Unit GP', 'Sales Total', 'Cost Total', 'Gross Profit', 'Salesperson']

MOVE_TYPE = {'out_invoice': 1, 'out_refund': -1}


class SalebyDateReport(models.TransientModel):
    _name = 'sale.date.report'
    _description = 'Generate a sales report between dates'

    date_from = fields.Date(string='Date from', default=date.today(), required=True)
    date_to = fields.Date(string='Date to', default=date.today(), required=True)
    route_id = fields.Many2one(comodel_name='crm.team', string='Route')
    partner_ids = fields.Many2many(comodel_name='res.partner', string='Customers')
    analysis = fields.Boolean(string='Analysis')

    def _prepare_by_date_data(self):
        load = []
        domain = [('state', '=', 'posted'), ('invoice_date', '>=', self.date_from),
                  ('invoice_date', '<=', self.date_to)]
        if self.route_id:
            domain += [('partner_id.team_id', '=', self.route_id.id)]

        payment_domain = [('state', '=', 'posted'), ('date', '>=', self.date_from),
                          ('date', '<=', self.date_to), ('payment_type', '=', 'inbound')]

        partners = self.partner_ids or self.env['res.partner'].search([])
        for partner in partners:
            invoice_domain = domain + [('move_type', '=', 'out_invoice'), ('partner_id', '=', partner.id)]
            refunds_domain = domain + [('move_type', '=', 'out_refund'), ('partner_id', '=', partner.id)]
            invoices = self.env['account.move'].search(invoice_domain)
            refunds = self.env['account.move'].search(refunds_domain)

            if not invoices and not refunds:
                continue
            sales = sum(invoices.mapped('amount_total')) - abs(sum(refunds.mapped('amount_total')))

            partner_payment_domain = payment_domain + [('partner_id', '=', partner.id)]
            payment_amount = sum(self.env['account.payment'].search(partner_payment_domain).mapped('amount'))

            load.append({
                'Customer': partner.name,
                'Route': partner.team_id and partner.team_id.name or '',
                'Sales': sales,
                'Paid': payment_amount,
                'Balance': sales - payment_amount,
                'Cumulative Bal': partner.total_due,
            })
        return load

    def action_by_date_generate(self):
        wb = xl.Workbook()
        ws = wb.active
        filename = 'Sales By Dates Report'
        customers = self.partner_ids and 'Select' or 'All'

        fr = 1  # current row
        ws[f'A{fr}'] = filename
        ws[f'A{fr}'].font = BIG_FONT
        ws[f'A{fr}'].alignment = CENTER_ALIGN
        ws.merge_cells(f'A{fr}:B{fr}')

        fr += 1
        ws[f'A{fr}'] = 'Print Out Date:'
        ws[f'B{fr}'] = datetime.now().strftime(DATETIME_FORMAT)
        fr += 1
        ws[f'A{fr}'] = 'Date:'
        ws[f'B{fr}'] = f'{self.date_from.strftime(DATE_FORMAT)} to {self.date_to.strftime(DATE_FORMAT)}'
        fr += 1
        ws[f'A{fr}'] = 'Customers:'
        ws[f'B{fr}'] = customers
        fr += 1
        ws[f'A{fr}'] = 'Currency:'
        ws[f'B{fr}'] = 'Balances in Home Currency'
        fr += 1
        ws[f'A{fr}'] = 'Route:'
        ws[f'B{fr}'] = self.route_id and self.route_id.name or 'All'
        fr += 2

        for row in range(2, 7):
            ws[f'A{row}'].font = MID_FONT
            ws[f'B{row}'].font = MID_FONT

        for index, col in enumerate(BY_DATE_COLS, start=1):
            letter = get_column_letter(index)
            ws[f'{letter}{fr}'] = col
            ws[f'{letter}{fr}'].font = MID_FONT

        fr += 1
        # POPULATING THE COLUMN VALUES
        start_data_row = fr
        for group in self._prepare_by_date_data():
            for index, col in enumerate(BY_DATE_COLS, start=1):
                letter = get_column_letter(index)
                ws[f'{letter}{fr}'] = group.get(col)
            fr += 1
        end_data_row = fr - 1

        # POPULATING THE TOTALS
        ws[f'A{fr}'] = 'Grand Total'
        ws[f'A{fr}'].font = MID_FONT

        for index, col in enumerate(BY_DATE_COLS[2:], start=3):
            letter = get_column_letter(index)
            cell = f'{letter}{fr}'
            col_range = f'{letter}{start_data_row}:{letter}{end_data_row}'
            ws[cell] = f'=SUM({col_range})'
            ws[cell].font = MID_FONT

        xlsx_path = self.env['excel.wizard'].create_xls()
        wb.save(xlsx_path)
        mes = f'Sales by date report generated by {self.env.user.name}.'
        datas = self.env['excel.wizard'].save_xls_file(xlsx_path, mes)

        attachment_id = self.env['excel.wizard'].create({
            'name': f'{filename}.xlsx',
            'report': datas
        })
        return {
            'name': 'Download Sales by Date Report',
            'context': self.env.context,
            'view_mode': 'form',
            'res_model': 'excel.wizard',
            'res_id': attachment_id.id,
            'type': 'ir.actions.act_window',
            'target': 'new'
        }

    def prepare_by_analysis_data(self):
        load = []
        domain = [('state', '=', 'posted'), ('move_type', 'in', ['out_invoice', 'out_refund']),
                  ('invoice_date', '>=', self.date_from), ('invoice_date', '<=', self.date_to)]
        if self.route_id:
            domain += [('partner_id.team_id', '=', self.route_id.id)]

        partners = self.partner_ids or self.env['res.partner'].search([])
        for partner in partners:
            invoice_domain = domain + [('partner_id', '=', partner.id)]
            invoices = self.env['account.move'].search(invoice_domain)

            if not invoices:
                continue
            invoice_lines = invoices.mapped(lambda i: (i.invoice_date, i.user_id, i.move_type, i.invoice_line_ids))

            for invoice_line in invoice_lines:
                invoice_date, salesperson, move_type, lines = invoice_line
                multiplier = 1 if move_type == 'out_invoice' else -1
                for line in lines.filtered(lambda p: p.product_id):
                    packaging_domain = [('product_id', '=', line.product_id.id), ('name', 'ilike', 'Crate')]
                    crates = self.env['product.packaging'].search(packaging_domain, limit=1).qty or 0
                    crates = crates and line.quantity / crates
                    load.append({
                        'Date': invoice_date,
                        'Dimension': partner.team_id and partner.team_id.name or '',
                        'Customer': partner.name,
                        'Product': line.product_id.name,
                        'QTY (pcs)': line.quantity * multiplier,
                        'QTY (crates)': float_round(crates, precision_rounding=0.01) * multiplier,
                        'Selling Price': line.price_unit * multiplier,
                        'Unit Cost': line.product_id.standard_price * multiplier,
                        'Unit GP': (line.price_unit - line.product_id.standard_price) * multiplier,
                        'Sales Total': line.price_unit * line.quantity * multiplier,
                        'Cost Total': line.product_id.standard_price * line.quantity * multiplier,
                        'Gross Profit': multiplier * ((line.price_unit * line.quantity) - (line.product_id.standard_price * line.quantity)),
                        'Salesperson': salesperson.name or '',
                    })
        return load

    def action_generate_analysis(self):
        wb = xl.Workbook()
        ws = wb.active
        filename = 'Sales Analysis Report'
        customers = self.partner_ids and 'Select' or 'All'

        fr = 1  # current row
        ws[f'A{fr}'] = filename
        ws[f'A{fr}'].font = BIG_FONT
        ws[f'A{fr}'].alignment = CENTER_ALIGN
        ws.merge_cells(f'A{fr}:B{fr}')

        fr += 1
        ws[f'A{fr}'] = 'Print Out Date:'
        ws[f'B{fr}'] = datetime.now().strftime(DATETIME_FORMAT)
        fr += 1
        ws[f'A{fr}'] = 'Date:'
        ws[f'B{fr}'] = f'{self.date_from.strftime(DATE_FORMAT)} to {self.date_to.strftime(DATE_FORMAT)}'
        fr += 1
        ws[f'A{fr}'] = 'Customers:'
        ws[f'B{fr}'] = customers
        fr += 1
        ws[f'A{fr}'] = 'Currency:'
        ws[f'B{fr}'] = 'Balances in Home Currency'
        fr += 1
        ws[f'A{fr}'] = 'Route:'
        ws[f'B{fr}'] = self.route_id and self.route_id.name or 'All'
        fr += 2

        for row in range(2, 7):
            ws[f'A{row}'].font = MID_FONT
            ws[f'B{row}'].font = MID_FONT

        for index, col in enumerate(ANALYSIS_COLS, start=1):
            letter = get_column_letter(index)
            ws[f'{letter}{fr}'] = col
            ws[f'{letter}{fr}'].font = MID_FONT

        fr += 1
        # POPULATING THE COLUMN VALUES
        for group in self.prepare_by_analysis_data():
            for index, col in enumerate(ANALYSIS_COLS, start=1):
                letter = get_column_letter(index)
                ws[f'{letter}{fr}'] = group.get(col)
            fr += 1

        xlsx_path = self.env['excel.wizard'].create_xls()
        wb.save(xlsx_path)
        mes = f'Sales analysis report generated by {self.env.user.name}.'
        datas = self.env['excel.wizard'].save_xls_file(xlsx_path, mes)

        attachment_id = self.env['excel.wizard'].create({
            'name': f'{filename}.xlsx',
            'report': datas
        })
        return {
            'name': 'Download Sales Analysis Report',
            'context': self.env.context,
            'view_mode': 'form',
            'res_model': 'excel.wizard',
            'res_id': attachment_id.id,
            'type': 'ir.actions.act_window',
            'target': 'new'
        }


class ReportDownload(models.TransientModel):
    _name = 'excel.wizard'
    _description = 'Download Excel Forms'

    name = fields.Char('File Name', size=64)
    report = fields.Binary('Excel Report', readonly=True)

    @api.model
    def create_xls(self):
        _, xls_path = tempfile.mkstemp(
            suffix='.xlsx', prefix='xlsreport.tmp.')
        return xls_path

    @api.model
    def save_xls_file(self, xls_path, message=None):
        if not message:
            message = f'A report has been generated by {self.env.user.name}'
        with open(xls_path, 'rb') as f:
            datas = base64.encodebytes(f.read())
            _logger.info(message)
            self.delete_tempfile(xls_path)
        return datas

    @api.model
    def delete_tempfile(self, path):
        try:
            os.unlink(path)
        except (OSError, IOError):
            _logger.error('Error when trying to remove file %s' % path)
